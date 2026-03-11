#!/usr/bin/env python3
"""
Refresh local analysis outputs after domain overrides change, without
refetching SERP data.
"""
import argparse
import json
import os

import openpyxl
import yaml

from classifiers import EntityClassifier
from generate_domain_override_candidates import (
    collect_candidates,
    load_overrides,
    normalize_domain,
    render_markdown,
)


def load_config_paths(config_path="config.yml"):
    config = {}
    if os.path.exists(config_path):
        with open(config_path, "r", encoding="utf-8") as f:
            config = yaml.safe_load(f) or {}
    files_cfg = config.get("files", {})
    return {
        "json": files_cfg.get("output_json", "market_analysis_v2.json"),
        "xlsx": files_cfg.get("output_xlsx", "market_analysis_v2.xlsx"),
        "overrides": files_cfg.get("domain_overrides", "domain_overrides.yml"),
        "candidates_report": "domain_override_candidates.md",
    }


def classify_domain_for_row(classifier, row):
    domain = normalize_domain(row.get("Link") or row.get("Source"))
    if not domain:
        return row.get("Entity_Type", "N/A")
    entity_type, _confidence, _evidence = classifier.classify(domain, None)
    return entity_type or "N/A"


def refresh_json(json_path, classifier):
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    updated = 0
    changed = 0
    for row in data.get("organic_results", []):
        new_type = classify_domain_for_row(classifier, row)
        old_type = row.get("Entity_Type", "N/A")
        row["Entity_Type"] = new_type
        updated += 1
        if old_type != new_type:
            changed += 1

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    return data, updated, changed


def refresh_xlsx(xlsx_path, classifier):
    if not os.path.exists(xlsx_path):
        return 0, 0

    wb = openpyxl.load_workbook(xlsx_path)
    if "Organic_Results" not in wb.sheetnames:
        return 0, 0

    ws = wb["Organic_Results"]
    headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    link_col = headers.get("Link")
    source_col = headers.get("Source")
    entity_col = headers.get("Entity_Type")
    if not link_col or not source_col or not entity_col:
        return 0, 0

    updated = 0
    changed = 0
    for row_idx in range(2, ws.max_row + 1):
        link = ws.cell(row=row_idx, column=link_col).value
        source = ws.cell(row=row_idx, column=source_col).value
        old_type = ws.cell(row=row_idx, column=entity_col).value or "N/A"
        new_type, _confidence, _evidence = classifier.classify(
            normalize_domain(link or source),
            None,
        )
        new_type = new_type or "N/A"
        ws.cell(row=row_idx, column=entity_col).value = new_type
        updated += 1
        if old_type != new_type:
            changed += 1

    wb.save(xlsx_path)
    return updated, changed


def regenerate_candidate_report(data, json_path, overrides_path, out_path):
    overrides = load_overrides(overrides_path)
    classifier = EntityClassifier(override_file=overrides_path)
    candidates = collect_candidates(
        data,
        overrides,
        classifier,
        min_rows=4,
        min_keywords=2,
    )
    report = render_markdown(candidates, json_path, overrides_path, 4, 2)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(report)
    return len(candidates)


def refresh_analysis_outputs(json_path, xlsx_path, overrides_path, candidates_report_path):
    classifier = EntityClassifier(override_file=overrides_path)
    data, json_updated, json_changed = refresh_json(json_path, classifier)
    xlsx_updated, xlsx_changed = refresh_xlsx(xlsx_path, classifier)
    candidate_count = regenerate_candidate_report(
        data,
        json_path,
        overrides_path,
        candidates_report_path,
    )
    return {
        "json_updated": json_updated,
        "json_changed": json_changed,
        "xlsx_updated": xlsx_updated,
        "xlsx_changed": xlsx_changed,
        "candidate_count": candidate_count,
    }


def main():
    defaults = load_config_paths()
    parser = argparse.ArgumentParser(
        description="Refresh local analysis files after domain override updates."
    )
    parser.add_argument("--json", default=defaults["json"])
    parser.add_argument("--xlsx", default=defaults["xlsx"])
    parser.add_argument("--overrides", default=defaults["overrides"])
    parser.add_argument("--candidate-report", default=defaults["candidates_report"])
    args = parser.parse_args()

    result = refresh_analysis_outputs(
        json_path=args.json,
        xlsx_path=args.xlsx,
        overrides_path=args.overrides,
        candidates_report_path=args.candidate_report,
    )
    print(
        f"Refreshed JSON rows: {result['json_updated']} (changed: {result['json_changed']})"
    )
    print(
        f"Refreshed XLSX rows: {result['xlsx_updated']} (changed: {result['xlsx_changed']})"
    )
    print(f"Regenerated candidate report. Remaining candidates: {result['candidate_count']}")


if __name__ == "__main__":
    main()
